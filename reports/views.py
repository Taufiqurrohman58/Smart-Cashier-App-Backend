from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import BasePermission
from django.db.models import Sum
from django.utils import timezone
from calendar import monthrange
from django.http import HttpResponse

from transactions.models import Transaction
from expenses.models import Expense

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from produks.models import KantinProduct
from transactions.models import TransactionItem


class IsAdmin(BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'admin'

@api_view(['GET'])
@permission_classes([IsAdmin])
def laporan_harian(request):
    date = request.query_params.get('date')
    if not date:
        return Response({'error': 'Parameter date wajib'}, status=400)

    transaksi = Transaction.objects.filter(created_at__date=date)
    pengeluaran = Expense.objects.filter(tanggal=date)

    detail_transaksi = []
    total_penjualan = 0

    for trx in transaksi:
        items = []
        for item in trx.items.all():
            items.append({
                'produk': item.product.name,
                'qty': item.qty,
                'subtotal': item.subtotal
            })

        total_penjualan += trx.total

        detail_transaksi.append({
            'invoice': trx.invoice,
            'kasir': trx.user.username,
            'items': items,
            'total': trx.total
        })

    total_pengeluaran = pengeluaran.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    return Response({
        'status': True,
        'periode': 'Harian',
        'tanggal': date,
        'penjualan': {
            'total_transaksi': transaksi.count(),
            'total_penjualan': total_penjualan,
            'detail': detail_transaksi
        },
        'pengeluaran': {
            'total_pengeluaran': total_pengeluaran,
            'detail': [
                {
                    'deskripsi': e.deskripsi,
                    'jumlah': e.jumlah,
                    'kasir': e.user.username
                } for e in pengeluaran
            ]
        },
        'ringkasan': {
            'total_penjualan': total_penjualan,
            'total_pengeluaran': total_pengeluaran,
            'laba_bersih': total_penjualan - total_pengeluaran
        }
    })

@api_view(['GET'])
@permission_classes([IsAdmin])
def laporan_bulanan(request):
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    if not month or not year:
        return Response({'error': 'month dan year wajib'}, status=400)

    transaksi = Transaction.objects.filter(
        created_at__month=month,
        created_at__year=year
    )

    pengeluaran = Expense.objects.filter(
        tanggal__month=month,
        tanggal__year=year
    )

    total_penjualan = transaksi.aggregate(
        total=Sum('total')
    )['total'] or 0

    total_pengeluaran = pengeluaran.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    return Response({
        'status': True,
        'periode': 'Bulanan',
        'bulan': month,
        'tahun': year,
        'penjualan': {
            'total_penjualan': total_penjualan,
            'total_transaksi': transaksi.count()
        },
        'pengeluaran': {
            'total_pengeluaran': total_pengeluaran
        },
        'ringkasan': {
            'laba_kotor': total_penjualan,
            'total_pengeluaran': total_pengeluaran,
            'laba_bersih': total_penjualan - total_pengeluaran
        }
    })

@api_view(['GET'])
@permission_classes([IsAdmin])
def laporan_tahunan(request):
    year = request.query_params.get('year')
    if not year:
        return Response({'error': 'year wajib'}, status=400)

    grafik = []

    for month in range(1, 13):
        penjualan = Transaction.objects.filter(
            created_at__year=year,
            created_at__month=month
        ).aggregate(total=Sum('total'))['total'] or 0

        pengeluaran = Expense.objects.filter(
            tanggal__year=year,
            tanggal__month=month
        ).aggregate(total=Sum('jumlah'))['total'] or 0

        grafik.append({
            'bulan': f"{month:02d}",
            'penjualan': penjualan,
            'pengeluaran': pengeluaran
        })

    total_penjualan = sum(g['penjualan'] for g in grafik)
    total_pengeluaran = sum(g['pengeluaran'] for g in grafik)

    return Response({
        'status': True,
        'periode': 'Tahunan',
        'tahun': year,
        'penjualan': {
            'total_penjualan': total_penjualan
        },
        'pengeluaran': {
            'total_pengeluaran': total_pengeluaran
        },
        'ringkasan': {
            'laba_bersih': total_penjualan - total_pengeluaran
        },
        'grafik': grafik
    })

@api_view(['GET'])
@permission_classes([IsAdmin])
def export_rekap_kantin_excel(request):
    date = request.query_params.get('date')
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    today = datetime.today()

    trx_filter = {}
    if date:
        trx_filter['transaction__created_at__date'] = date
        periode = date
    elif month and year:
        trx_filter['transaction__created_at__month'] = month
        trx_filter['transaction__created_at__year'] = year
        periode = f"{month}-{year}"
    elif year:
        trx_filter['transaction__created_at__year'] = year
        periode = year
    else:
        trx_filter['transaction__created_at__date'] = today.date()
        periode = today.strftime('%Y-%m-%d')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rekap Stok Kantin'

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'REKAP STOK KANTIN ({periode})'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = [
        'No',
        'Nama Barang',
        'Stok Masuk',
        'Stok Awal',
        'Jumlah Stok',
        'Satuan',
        'Terjual',
        'Sisa Stok',
        'Harga'
    ]

    ws.append(headers)

    header_row = 2
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    kantin_products = KantinProduct.objects.select_related('product_gudang')

    row = 3
    no = 1
    for kp in kantin_products:
        terjual = TransactionItem.objects.filter(
            product=kp,
            **trx_filter
        ).aggregate(total=Sum('qty'))['total'] or 0

        stok_awal = kp.stock_kantin + terjual
        stok_masuk = 0
        jumlah_stok = stok_awal + stok_masuk

        data = [
            no,
            kp.product_gudang.name,
            stok_masuk,
            stok_awal,
            jumlah_stok,
            kp.product_gudang.satuan,
            terjual,
            kp.stock_kantin,
            kp.product_gudang.price
        ]

        ws.append(data)

        for col in range(1, len(data) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in [1, 3, 4, 5, 7, 8, 9]:
                cell.alignment = center_align

        row += 1
        no += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=rekap_kantin_{periode}.xlsx'
    )

    wb.save(response)
    return response
